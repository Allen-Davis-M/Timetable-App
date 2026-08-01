"""
Export a generated timetable to Excel (.xlsx) or PDF.

Both formats show the same two views already in the frontend
(TimetableTab.jsx): one grid per class-group section ("By Section") and
one grid per teacher ("By Teacher", their combined schedule across every
section they teach) — so what an admin sees on screen, and what they hand
out to a section or a teacher on paper, always match.

Shared data-gathering happens once in `_load_grid_data` so the Excel and
PDF builders can never disagree about what's actually in the timetable —
they just render the same grid data two different ways.
"""
import io
import re
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from app.models.school import ClassGroup, Period, Room, Subject, Teacher, Timetable

DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

HEADER_FILL = "1E293B"  # slate-800, matches the frontend's dark header rows


def _load_grid_data(db: Session, timetable: Timetable) -> dict:
    school_id = timetable.school_id
    periods = db.query(Period).filter(Period.school_id == school_id).all()
    days = sorted({p.day_of_week for p in periods})
    orders = sorted({p.order for p in periods})
    period_by_day_order = {(p.day_of_week, p.order): p for p in periods}

    class_groups = db.query(ClassGroup).filter(ClassGroup.school_id == school_id).all()
    teachers = db.query(Teacher).filter(Teacher.school_id == school_id).all()
    subjects_by_id = {s.id: s for s in db.query(Subject).filter(Subject.school_id == school_id).all()}
    teachers_by_id = {t.id: t for t in teachers}
    class_groups_by_id = {c.id: c for c in class_groups}
    rooms_by_id = {r.id: r for r in db.query(Room).filter(Room.school_id == school_id).all()}

    entries_by_class_period = defaultdict(list)
    entries_by_teacher_period = defaultdict(list)
    for e in timetable.entries:
        entries_by_class_period[(e.class_group_id, e.period_id)].append(e)
        entries_by_teacher_period[(e.teacher_id, e.period_id)].append(e)

    return {
        "days": days,
        "orders": orders,
        "period_by_day_order": period_by_day_order,
        "class_groups": class_groups,
        "teachers": teachers,
        "subjects_by_id": subjects_by_id,
        "teachers_by_id": teachers_by_id,
        "class_groups_by_id": class_groups_by_id,
        "rooms_by_id": rooms_by_id,
        "entries_by_class_period": entries_by_class_period,
        "entries_by_teacher_period": entries_by_teacher_period,
    }


def _class_group_label(cg: ClassGroup) -> str:
    return f"{cg.grade} - {cg.name}" if cg.grade else cg.name


def _cell_text(entry, data: dict, mode: str) -> str:
    """mode='section': shows subject + teacher (for a section's own grid).
    mode='teacher': shows subject + section (for a teacher's own grid).
    Both also show the assigned room, if any (see solver.py's best-effort
    room assignment — not every entry is guaranteed a room)."""
    subject = data["subjects_by_id"].get(entry.subject_id)
    subject_name = subject.name if subject else "?"
    room = data["rooms_by_id"].get(entry.room_id) if entry.room_id else None
    room_suffix = f"\n[{room.name}]" if room else ""
    if mode == "section":
        teacher = data["teachers_by_id"].get(entry.teacher_id)
        return f"{subject_name}\n({teacher.name if teacher else '?'}){room_suffix}"
    cg = data["class_groups_by_id"].get(entry.class_group_id)
    return f"{subject_name}\n({_class_group_label(cg) if cg else '?'}){room_suffix}"


def _grid_rows(data: dict, entries_index: dict, entry_key_prefix, mode: str) -> list[list[str]]:
    """Builds a list of rows (including the header row) for one grid:
    row[0] = "Period N", followed by one cell per day. `entries_index` is
    either entries_by_class_period or entries_by_teacher_period;
    entry_key_prefix is the class_group_id or teacher_id to look up."""
    rows = [["Period"] + [DAY_NAMES[d] for d in data["days"]]]
    for order in data["orders"]:
        row = [f"Period {order + 1}"]
        for day in data["days"]:
            period = data["period_by_day_order"].get((day, order))
            text = ""
            if period:
                entries = entries_index.get((entry_key_prefix, period.id), [])
                text = "\n".join(_cell_text(e, data, mode) for e in entries)
            row.append(text)
        rows.append(row)
    return rows


def _safe_sheet_name(name: str) -> str:
    """Excel sheet names can't contain : \\ / ? * [ ] and are capped at 31 chars."""
    cleaned = re.sub(r"[:\\/?*\[\]]", "-", name)
    return cleaned[:31] or "Sheet"


def build_excel(db: Session, timetable: Timetable) -> bytes:
    data = _load_grid_data(db, timetable)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    title_font = Font(bold=True, size=14)
    wrap = Alignment(wrap_text=True, vertical="top", horizontal="left")

    def write_sheet(title: str, rows: list[list[str]]):
        ws = wb.create_sheet(_safe_sheet_name(title))
        ws.append([title])
        ws["A1"].font = title_font
        ws.append([])
        for row in rows:
            ws.append(row)
        header_row_num = 3
        for col in range(1, len(rows[0]) + 1):
            cell = ws.cell(row=header_row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = 24
        for row in ws.iter_rows(min_row=header_row_num + 1, max_row=ws.max_row):
            for cell in row:
                cell.alignment = wrap
            ws.row_dimensions[row[0].row].height = 42

    for cg in data["class_groups"]:
        rows = _grid_rows(data, data["entries_by_class_period"], cg.id, mode="section")
        if any(any(r[1:]) for r in rows[1:]):  # skip sections with nothing scheduled
            write_sheet(f"{_class_group_label(cg)}", rows)

    for teacher in data["teachers"]:
        rows = _grid_rows(data, data["entries_by_teacher_period"], teacher.id, mode="teacher")
        if any(any(r[1:]) for r in rows[1:]):
            write_sheet(f"{teacher.name}", rows)

    if not wb.sheetnames:
        wb.create_sheet("Timetable").append(["This timetable has no entries yet."])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf(db: Session, timetable: Timetable) -> bytes:
    data = _load_grid_data(db, timetable)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TimetableTitle", parent=styles["Heading1"], spaceAfter=10)
    cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=8, leading=10)

    story = []
    n_cols = len(data["days"]) + 1
    col_width = (landscape(A4)[0] - 2.4 * cm) / n_cols

    def add_table(title: str, rows: list[list[str]]):
        story.append(Paragraph(title, title_style))
        # Wrap every non-header cell's text in a Paragraph so long entries
        # wrap instead of overflowing the column.
        wrapped_rows = [rows[0]] + [
            [row[0]] + [Paragraph(cell.replace("\n", "<br/>"), cell_style) if cell else "" for cell in row[1:]]
            for row in rows[1:]
        ]
        table = Table(wrapped_rows, colWidths=[col_width] * n_cols, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{HEADER_FILL}")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 1), (0, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ]))
        story.append(table)
        story.append(Spacer(1, 0.6 * cm))

    sections_added = 0
    for cg in data["class_groups"]:
        rows = _grid_rows(data, data["entries_by_class_period"], cg.id, mode="section")
        if any(any(r[1:]) for r in rows[1:]):
            if sections_added:
                story.append(PageBreak())
            add_table(f"{_class_group_label(cg)} — Timetable", rows)
            sections_added += 1

    teachers_added = 0
    for teacher in data["teachers"]:
        rows = _grid_rows(data, data["entries_by_teacher_period"], teacher.id, mode="teacher")
        if any(any(r[1:]) for r in rows[1:]):
            story.append(PageBreak())
            add_table(f"{teacher.name} — Teaching Schedule", rows)
            teachers_added += 1

    if not story:
        story.append(Paragraph("This timetable has no entries yet.", styles["Normal"]))

    doc.build(story)
    return buf.getvalue()
